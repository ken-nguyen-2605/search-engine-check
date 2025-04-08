from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import TimeoutException
import time
from questions_data import question_tuple

def search_virtual_treasury(phrases_and_entities):
    """
    Search phrases on virtualtreasury.ie and check for entities in specific sections.
    
    Args:
        phrases_and_entities: List of tuples (search_phrase, entity_to_find)
    """
    base_url = "https://virtualtreasury.ie"
    
    # Configure Chrome options
    chrome_options = Options()
    # Uncomment for headless mode if needed
    # chrome_options.add_argument("--headless=new")
    chrome_options.add_argument("--window-size=1920,1080")
    chrome_options.add_argument("--disable-notifications")
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--no-sandbox")
    
    # Initialize the driver
    driver = webdriver.Chrome(options=chrome_options)
    
    results = []
    
    try:
        for phrase, entity in phrases_and_entities:
            print(f"\nSearching for: {phrase}")
            
            # Use the search URL format directly instead of using the form
            encoded_phrase = phrase.replace(" ", "%20")
            search_url = f"{base_url}/search-results?totalElementsInt=10&kwOperList=ANY&kwList={encoded_phrase}&kwSearchFieldList=all&resultSorting=relevance&pageNumberInt=0"
            
            driver.get(search_url)
            time.sleep(5)  # Give the page time to load
            
            # Check if results are present
            try:
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CLASS_NAME, "spg-result"))
                )
                result_divs = driver.find_elements(By.CLASS_NAME, "spg-result")
                print(f"Found {len(result_divs)} results")
            except TimeoutException:
                print("No results found or timeout waiting for results!")
                continue
            
            if not result_divs:
                print("No results found!")
                continue
            
            # Process each result
            for idx, div in enumerate(result_divs, 1):
                try:
                    # Find the link in the result (using your HTML structure)
                    link_elem = div.find_element(By.CSS_SELECTOR, "a[target='_blank'][href]")
                    full_url = link_elem.get_attribute('href')
                    print(f"Processing result {idx}: {full_url}")
                    
                    # Open the result in a new tab
                    driver.execute_script("window.open('');")
                    driver.switch_to.window(driver.window_handles[1])
                    driver.get(full_url)
                    time.sleep(3)  # Wait for page to load
                    
                    # Look for the two sections
                    section1_text = ""
                    section2_text = ""
                    
                    try:
                        section1 = WebDriverWait(driver, 10).until(
                            EC.presence_of_element_located((By.CSS_SELECTOR, "section.module-page-title"))
                        )
                        section1_text = section1.text
                    except TimeoutException:
                        print("Section 1 not found")
                        
                    try:
                        section2 = WebDriverWait(driver, 10).until(
                            EC.presence_of_element_located((By.CSS_SELECTOR, "section.module-item"))
                        )
                        section2_text = section2.text
                    except TimeoutException:
                        print("Section 2 not found")
                    
                    # Combine text and check for entity
                    combined_text = (section1_text + " " + section2_text).lower()
                    entity_lower = entity.lower()
                    entity_words = entity_lower.split()
                    
                    if entity_lower in combined_text:
                        print("y")
                        results.append(("y", full_url))
                    else:
                        # Check which words from entity exist
                        words_found = [word for word in entity_words if word.lower() in combined_text]
                        
                        if len(words_found) == 0:
                            reason = "No words from the entity found"
                        elif len(words_found) == 1:
                            reason = f"Exists only 1 word in the entity ({words_found[0]})"
                        elif 1 < len(words_found) < len(entity_words):
                            words_str = ", ".join(words_found)
                            reason = f"Exists only {len(words_found)} words in the entity ({words_str})"
                        elif len(words_found) == len(entity_words):
                            reason = "There exists all of the words but in various places"
                        else:
                            reason = "Another case"
                        
                        print(f"Link: {full_url}")
                        print(f"Reason: {reason}")
                        results.append((reason, full_url))
                    
                    # Close tab and switch back to results
                    driver.close()
                    driver.switch_to.window(driver.window_handles[0])
                    
                except Exception as e:
                    print(f"Error processing result {idx}: {str(e)}")
                    # Make sure we return to the main window
                    if len(driver.window_handles) > 1:
                        driver.close()
                        driver.switch_to.window(driver.window_handles[0])
    
    finally:
        driver.quit()
    
    return results

def save_results_to_excel(search_items, results, filename="search_results.xlsx"):
    """
    Save search results to an Excel file.
    
    Args:
        search_items: List of tuples (search_phrase, entity_to_find)
        results: List of tuples (reason, url) from search_virtual_treasury
        filename: Name of the Excel file to save
        
    Note:
        This function assumes results are ordered by search phrase,
        with each search phrase having a variable number of results.
    """
    import openpyxl
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter
    
    # Create a new workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Add headers
    headers = ['No', 'Query'] + [f'Doc{i}' for i in range(1, 11)]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    
    # Collect results for each search phrase
    grouped_results = []
    current_result_idx = 0
    
    for phrase, entity in search_items:
        # Get results for this search phrase (up to 10)
        phrase_results = []
        for _ in range(10):  # Up to 10 results per search phrase
            if current_result_idx < len(results):
                phrase_results.append(results[current_result_idx])
                current_result_idx += 1
            else:
                break
        
        grouped_results.append(phrase_results)
    
    # Add data to the Excel sheet
    for idx, ((phrase, entity), phrase_results) in enumerate(zip(search_items, grouped_results), 1):
        row_idx = idx + 1  # Row index in the Excel sheet (accounting for headers)
        
        # Add No and Query
        ws.cell(row=row_idx, column=1, value=idx)       # No
        ws.cell(row=row_idx, column=2, value=phrase)    # Query
        
        # Add results for this search phrase
        for doc_idx, (reason, url) in enumerate(phrase_results, 1):
            col_idx = doc_idx + 2  # Doc columns start at column 3
            
            if reason == "y":
                cell_value = f"y\nLink: {url}"
            else:
                cell_value = f"Link: {url}\nReason: {reason}"
            
            ws.cell(row=row_idx, column=col_idx, value=cell_value)
            ws.cell(row=row_idx, column=col_idx).alignment = Alignment(wrap_text=True)
    
    # Set column widths
    ws.column_dimensions['A'].width = 5   # No
    ws.column_dimensions['B'].width = 30  # Query
    for col_idx in range(3, 13):  # Doc1 to Doc10
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 25
    
    # Save the workbook
    wb.save(filename)
    print(f"Results saved to {filename}")

def main():
    # Example list of phrases and entities to search for
    search_items = question_tuple
    
    results = search_virtual_treasury(search_items)
    
    print("\nSearch complete!")
    
    # Save results to Excel
    save_results_to_excel(search_items, results, "virtual_treasury_search_results.xlsx")

if __name__ == "__main__":
    main()